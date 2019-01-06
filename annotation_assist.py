from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import sys
import os
import os.path
import time
import re
import datetime


class AnnotationAssist(object):

    def __init__(self):
        self.__data_path = 'D:\\dev\\src\\work\\brat\\cailianpress'
        self.__root_dir = 'D:\\dev\\src\\work\\brat\\cailianpress\\split'

    def get_all_a_stocks_name_from(self, path):
        if os.path.exists(path):
            if os.access(path, os.R_OK):
                work_book = load_workbook(filename=path)
                ws = work_book.get_active_sheet()

                stock_file_path = os.path.join(self.__data_path, '产品_股票.txt')
                try:
                    with open(file=stock_file_path, mode='w+', encoding='utf-8') as file:
                        for r in range(2, 3570):
                            stock = str(ws.cell(row=r, column=1).value).split('.')[0] + "，" + \
                                    str(ws.cell(row=r, column=2).value).replace('*', '') + '\n'
                            file.write(stock)
                except Exception as e:
                    print(str(e))

                work_book.close()
            else:
                print(path + " 不可读！")
        else:
            print(path + " 不存在！")

    def load_keywords_list_from(self, path):
        if os.path.exists(path):
            if os.access(path, os.R_OK):
                lines = None
                keywords = []
                try:
                    with open(file=path, mode='r', encoding='utf-8') as file:
                        lines = file.readlines()
                        print(path + "== 已读取行数:" + str(len(lines)))
                except Exception as e:
                    print(str(e))
                entity_type = path.split('\\')[-1].replace('.txt', '')
                print(entity_type)
                if not lines:
                    return None
                for line in lines:
                    if not line:
                        continue
                    words = line.split('，')
                    for word in words:
                        word = word.strip().replace('\n', '')
                        if ',' in word:
                            print(path + ' 发现英文逗号:' + word)
                        entity_pair = [word, entity_type]
                        #print(entity_pair[0] + '-' + entity_pair[1])
                        keywords.append(entity_pair)
                return keywords
            else:
                print(path + " 不可读！")
                return None
        else:
            print(path + " 不存在！")
            return None

    def load_all_keywords(self, path):
        keyword_list = []
        for root, dirs, files in os.walk(path):
            txt_pattern = re.compile(r'.*txt$')
            for file in files:
                if txt_pattern.match(file):
                    txt_file_path = os.path.join(root, file)
                    words = self.load_keywords_list_from(txt_file_path)
                    if not words:
                        continue
                    keyword_list.extend(words)
        idx = 0
        sorted_keyword_list = sorted(keyword_list, key=lambda x: (len(x[0])), reverse=True)
        '''
        for key in sorted_keyword_list:
            idx += 1
            if idx % 5 == 0:
                print(key[0] + '-' + key[1])
            else:
                print(key[0] + '-' + key[1] + '，', end='')
        '''
        return sorted_keyword_list

    # matched_keyword structure: [keyword, start, end]
    def find_new_entity_record(self, ann_file_path, matched_keyword_list, entity_pattern, entity_type, record_index):
        new_entity_records_list = []
        try:
            with open(file=ann_file_path, mode='r', encoding='utf-8') as af:
                lines = af.readlines()
                if len(matched_keyword_list) > 0:
                    for mkl in matched_keyword_list:
                        is_founded = False
                        for line in lines:
                            entity_match = entity_pattern.match(line)
                            if entity_match:
                                start = int(entity_match.group(1))
                                end = int(entity_match.group(2))
                                mkl1 = int(mkl[1])  # start index
                                mkl2 = int(mkl[2])  # end index
                                if mkl1 == start:
                                    is_founded = True
                                    if end == mkl2:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 已存在')
                                        break
                                    elif end > mkl2:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 被包含匹配')
                                        break
                                    else:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 包含匹配')
                                        break
                                elif end > mkl1 > start:
                                    is_founded = True
                                    print(mkl[0] + ', ' + mkl[1] + ', '
                                          + mkl[2] + ' 内包含匹配')
                                    break
                                elif end > mkl2 > start:
                                    is_founded = True
                                    print(mkl[0] + ', ' + mkl[1] + ', '
                                          + mkl[2] + ' 交叉匹配')
                                    break
                                else:
                                    continue

                        if not is_founded:
                            record_index += 1
                            record = 'T' + str(record_index) + '\t' + entity_type + ' ' + \
                                     str(mkl[1]) + ' ' + str(mkl[2]) + '\t' + mkl[0] + '\n'
                            #print(record)
                            new_entity_records_list.append(record)
        except Exception as e:
            print(str(e))
        return new_entity_records_list

    def find_new_event_record(self, ann_file_path, matched_keyword_list,
                              entity_pattern, event_type,
                              record_index_of_entity,
                              record_index_of_event):
        new_event_records_list = []
        try:
            with open(file=ann_file_path, mode='r', encoding='utf-8') as af:
                lines = af.readlines()
                if len(matched_keyword_list) > 0:
                    for mkl in matched_keyword_list:
                        is_founded = False
                        for line in lines:
                            event_match = entity_pattern.match(line)
                            if event_match:
                                start = int(event_match.group(1))
                                end = int(event_match.group(2))
                                mkl1 = int(mkl[1])  # start index
                                mkl2 = int(mkl[2])  # end index
                                if mkl1 == start:
                                    is_founded = True
                                    if end == mkl2:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 已存在')
                                        break
                                    elif end > mkl2:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 被包含匹配')
                                        break
                                    else:
                                        print(mkl[0] + ', ' + mkl[1] + ', ' +
                                              mkl[2] + ' 包含匹配')
                                        break
                                elif end > mkl1 > start:
                                    is_founded = True
                                    print(mkl[0] + ', ' + mkl[1] + ', '
                                          + mkl[2] + ' 内包含匹配')
                                    break
                                elif end > mkl2 > start:
                                    is_founded = True
                                    print(mkl[0] + ', ' + mkl[1] + ', '
                                          + mkl[2] + ' 交叉匹配')
                                    break
                                else:
                                    continue

                        if not is_founded:
                            record_index_of_entity += 1
                            record_index_of_event += 1
                            entity_record = 'T' + str(record_index_of_entity) + '\t' + event_type + ' ' + \
                                            str(mkl[1]) + ' ' + str(mkl[2]) + '\t' + mkl[0] + '\n'
                            event_record = 'E' + str(record_index_of_event) + '\t' + event_type + ':T' + \
                                           str(record_index_of_entity) + ' ' + '\n'
                            record_pairs = [entity_record, event_record]
                            #print(record)
                            new_event_records_list.append(record_pairs)
        except Exception as e:
            print(str(e))
        return new_event_records_list

    def write_event_record_list_to_ann_file(self, key_list, ann_file_path):
        try:
            with open(file=ann_file_path, mode='a+', encoding='utf-8') as f:
                for key in key_list:
                    print(key[0])
                    print(key[1])
                    #f.write(key[0])
                    #f.write(key[1])
        except Exception as e:
            print(str(e))

    def write_entity_record_list_to_ann_file(self, key_list, ann_file_path):
        try:
            with open(file=ann_file_path, mode='a+', encoding='utf-8') as f:
                for key in key_list:
                    print(key)
                    #f.write(key)
        except Exception as e:
            print(str(e))

    def process_entity_record_annotation(self, entity_pattern, text, entity_type, record_idx, ann_file_path):
        matches = re.finditer(entity_pattern, text)
        matched_list = []
        for m in matches:
            matched_key_info = [m.group(), str(m.start()), str(m.end())]
            matched_list.append(matched_key_info)

        common_entity_pattern = re.compile(r'T\d*\t\S* (\d*) (\d*)\t.*')
        new_list = self.find_new_entity_record(ann_file_path,
                                               matched_list,
                                               common_entity_pattern,
                                               entity_type,
                                               record_idx)
        if len(new_list):
            print('record after ' + entity_type + ': ' + str(record_idx + len(new_list)))
        self.write_entity_record_list_to_ann_file(new_list, ann_file_path)
        return len(new_list)

    def auto_annotation(self, keywords_of_entity, keywords_of_event):
        txt_pattern = re.compile(r'.*txt$')
        ann_pattern = re.compile(r'.*ann$')
        entity_pattern = re.compile(r'T\d*\t\S* (\d*) (\d*)\t.*')
        # to match a period, such as 自2018年1月1日起，2005年期间，本月1号起
        period_pattern = re.compile(
            r'(((去年|今年|\d{2,4}年)?(\d{1,2}|本|上)月(\d{1,2}(日|号))?)|((去|今|\d{2,4})年)|(\d{1,2}(日|号)))(以来|开始|始|起|期间)')
        # to match a conference, such as, 2017年第10界中国各种博览会, 第10界中国各种大会
        conference_pattern = re.compile(
            r'((\d{2,4}年)|(第\S*界))(第\S界)?\S*?(大会|会议|博览会|展览会|讨论会|研讨会|展销会|小组会|年会|论坛|峰会)')
        # to match a date
        date_pattern = re.compile(
            r'((((去|今|\d{2,4})年)?(\d{1,2}|本|上)月(\d{1,2}(日|号))?)|(\d{1,2}(日|号)))(上午|下午|中午|晚间|夜晚|凌晨)?')
        # to match a stock index, percentage number
        stock_index_pattern = re.compile(r'\d+(\.\d+)?(点关口|关口|点大关|大关|个百分点|个点|点)')
        # to match a percentage number
        percent_pattern = re.compile(r'\d+(\.\d+)?%')
        # to match a price
        price_pattern = re.compile(
            r'\d+(\.\d+)?(万元/平方米|美元/桶|美元/吨|美金/吨|元/立方米|万桶/日|欧元/公斤|6元/手|元/吨|元/股|美元关口|亿元|万美元|万亿美元|美元|亿美元|亿英镑|万英镑|英镑|加元|亿加元|元人民币|万亿|美金|亿日元|千万日元|日元|万元|元)')
        # to match a measure
        measure_pattern = re.compile(r'\d+(\.\d+)?(千公顷|万公顷|公顷|万平方米|千平方米|平方米)')
        # to match a count
        count_pattern = re.compile(
            r'第?\d+(起|次|手|宗|份|辆|台|套|场|件|项|家|万股|千股|股)')  # 个 usually used with 月， they stand for a period
        for root, dirs, files in os.walk(self.__root_dir):
            for file in files:
                if txt_pattern.match(file):
                    txt_file_path = os.path.join(root, file)
                    print('current processing : ' + txt_file_path)
                    text = ''
                    try:
                        with open(file=txt_file_path, mode='r', encoding='utf-8') as f:
                            text = f.readline()
                            text = text.strip('\n')
                            print(text)
                    except Exception as e:
                        print(str(e))

                    ann_file_path = txt_file_path.replace('.txt', '.ann')
                    entity_record_idx = self.count_of_entities(ann_file_path)
                    event_record_idx = self.count_of_events(ann_file_path)
                    # matched info structure : [key, start, end]

                    # to check period keywords
                    entity_record_idx += self.process_entity_record_annotation(period_pattern,
                                                                               text,
                                                                               r'时间_时期',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check conference keywords
                    entity_record_idx += self.process_entity_record_annotation(conference_pattern,
                                                                               text,
                                                                               r'事件_会议',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check date keywords
                    entity_record_idx += self.process_entity_record_annotation(date_pattern,
                                                                               text,
                                                                               r'时间_日期',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check stock index keywords
                    entity_record_idx += self.process_entity_record_annotation(stock_index_pattern,
                                                                               text,
                                                                               r'数字_经金类指数',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check percentage keywords
                    entity_record_idx += self.process_entity_record_annotation(percent_pattern,
                                                                               text,
                                                                               r'数字_百分比',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check price keywords
                    entity_record_idx += self.process_entity_record_annotation(price_pattern,
                                                                               text,
                                                                               r'数字_钱',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check measure keywords
                    entity_record_idx += self.process_entity_record_annotation(measure_pattern,
                                                                               text,
                                                                               r'数字_测量',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check count keywords
                    entity_record_idx += self.process_entity_record_annotation(count_pattern,
                                                                               text,
                                                                               r'数字_计数',
                                                                               entity_record_idx,
                                                                               ann_file_path)
                    # to check all the entity keywords
                    for kw in keywords_of_entity:
                        # a keyword could exist in a news more than one times
                        # keyword structure: [word, entity_type]
                        entity_keyword_pattern = re.compile(kw[0])
                        entity_keyword_match_iter = re.finditer(entity_keyword_pattern, text)
                        matched_entity_keyword_list = []  # [matched_kw_info, matched_kw_info, matched_kw_info]
                        for en_it in entity_keyword_match_iter:
                            # mkl structs [keyword, start, end]
                            matched_kw_info = [kw[0], str(en_it.start()), str(en_it.end())]
                            print('--' + matched_kw_info[0] + ',' + matched_kw_info[1] + ',' + matched_kw_info[2])
                            matched_entity_keyword_list.append(matched_kw_info)

                        # to check these keywords whether exists in annotation file
                        new_entity_records_list = self.find_new_entity_record(ann_file_path,
                                                                              matched_entity_keyword_list,
                                                                              entity_pattern,
                                                                              kw[1],
                                                                              entity_record_idx)
                        entity_record_idx += len(new_entity_records_list)

                        if len(new_entity_records_list) > 0:
                            try:
                                with open(file=ann_file_path, mode='a+', encoding='utf-8') as af:
                                    for record in new_entity_records_list:
                                        print(record)
                                        #af.write(record)
                            except Exception as e:
                                print(str(e))

                    # to process event keywords
                    for kw in keywords_of_event:
                        event_keyword_pattern = re.compile(kw[0])
                        event_keyword_match_iter = re.finditer(event_keyword_pattern, text)
                        matched_event_keyword_list = []  # [matched_kw_info, matched_kw_info, matched_kw_info]
                        for ev_it in event_keyword_match_iter:
                            # mkl structs [keyword, start, end]
                            matched_ekw_info = [kw[0], str(ev_it.start()), str(ev_it.end())]
                            print('--' + matched_ekw_info[0] + ',' + matched_ekw_info[1] + ',' + matched_ekw_info[2])
                            matched_event_keyword_list.append(matched_ekw_info)

                        new_event_record_list = self.find_new_event_record(ann_file_path,
                                                                           matched_event_keyword_list,
                                                                           entity_pattern,
                                                                           kw[1],
                                                                           entity_record_idx,
                                                                           event_record_idx)
                        entity_record_idx += len(new_event_record_list)
                        event_record_idx += len(new_event_record_list)
                        if len(new_event_record_list) > 0:
                            self.write_event_record_list_to_ann_file(new_event_record_list, ann_file_path)
                elif ann_pattern.match(file):
                    #pass
                    file_path = os.path.join(root, file)
                else:
                    print('遇到非 .txt, .ann文件了，跳过')

    def count_of_entities(self, path):
        # entity_pattern = re.compile(r'T(\d*)\t(\S*) (\d*) (\d*)\t(\S*)\n')
        entity_pattern = re.compile(r'T(\d+)')
        count = 0
        try:
            with open(file=path, mode='r', encoding='utf-8') as file:
                lines = file.readlines()
                for line in lines:
                    entity_match = entity_pattern.match(line)
                    if entity_match:
                        count += 1
        except Exception as e:
            print(str(e))

        return count

    def count_of_events(self, path):
        event_pattern = re.compile(r'E(\d+)')
        count = 0
        try:
            with open(file=path, mode='r', encoding='utf-8') as file:
                lines = file.readlines()
                for line in lines:
                    entity_match = event_pattern.match(line)
                    if entity_match:
                        count += 1
        except Exception as e:
            print(str(e))

        return count

    '''
    these code could only check a word in one time
    so, do not use it anymore
    '''
    def search_exists_span_in_annotation_file(self, start_idx, path):
        entity_pattern = re.compile(r'T\d*\t\S* (\d*) (\d*)\t.*')
        try:
            with open(file=path, mode='r', encoding='utf-8') as file:
                lines = file.readlines()
                for line in lines:
                    entity_match = entity_pattern.match(line)
                    if entity_match:
                        start = entity_match.group(1)
                        if int(start_idx) == int(start):
                            end = entity_match.group(2)
                            return start, end
                        else:
                            pass
        except Exception as e:
            print(str(e))
        return -1, -1

    @staticmethod
    def print_keywords(key_words):
        idx = 0
        for keyword in key_words:
            idx += 1
            if idx % 10 == 0:
                print(str(keyword[0]) + '-' + str(keyword[1]))
            else:
                print(str(keyword[0]) + '-' + str(keyword[1]) + '，', end='')

    @staticmethod
    def print_all_current_news(path):
        txt_pattern = re.compile(r'.*txt$')
        lines = []
        for root, dirs, files in os.walk(path):
            for file in files:
                if txt_pattern.match(file):
                    txt_file_path = os.path.join(root, file)
                    try:
                        with open(file=txt_file_path, mode='r', encoding='utf-8') as file:
                            line = file.readline()
                            lines.append(line)
                    except Exception as e:
                        print(str(e))
        for line in lines:
            print(line)


aa = AnnotationAssist()
#aa.get_all_a_stocks_name_from('D:\\dev\\src\\work\\brat\cailianpress\\全部A股.xlsx')
#keywords = aa.load_keywords_list_from("D:\\dev\\src\\work\\brat\cailianpress\\中国人口中心.txt")
#aa.print_keywords(keywords)
#aa.auto_annotation('', keywords)
#aa.print_all_current_news('D:\\dev\\src\\work\\brat\\cailianpress\\split')
keyword_of_entity = aa.load_all_keywords(r'D:\dev\src\work\brat\cailianpress\entity_keywords')
keyword_of_event = aa.load_all_keywords(r'D:\dev\src\work\brat\cailianpress\event_keywords')
aa.auto_annotation(keyword_of_entity, keyword_of_event)
#count = aa.count_of_entities(r'D:\dev\src\work\brat\cailianpress\split\cailianpress_2018-05-31\2018-05-31_02_27_457.ann')
#print(count)
#keywords = aa.load_keywords_list_from(r'D:\dev\src\work\brat\cailianpress\keywords\地缘政治实体_人口中心.txt')
#aa.auto_annotation_of_entity('地缘政治实体_人口中心', keywords)
#aa.print_keywords(keyword_of_entity)
#aa.print_keywords(keyword_of_event)